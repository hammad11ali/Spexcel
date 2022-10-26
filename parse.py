
import sys
import openpyxl
# from openpyxl.cell import get_column_letter
from openpyxl.utils import get_column_letter
import time
import os

os.system('cls')

print("\n\n***Welcome to Spool to Excel Parser and Converter***\n\n")

if len(sys.argv) < 2:
    print('Please provide spool file name with extension')
    print('Format: spexcel filename')
    sys.exit()
try:
    file = open(sys.argv[1], 'r')
except:
    print('File not found')
    sys.exit()

print('File opened')
print('Reading and Parsing file data')
new = False
data = False

query = ""
header = ""
rows = []
final = []
once = False
read = False
while True:
    line = file.readline()
    if not line:
        break
    if line.startswith('SQL>'):
        new = True
        data = False
        if rows:
            rows.remove(rows[0])
        final.append((query, header, rows))
        query = ""
        header = ""
        rows = []
    if new:
        query += ' '+line
        if line == "\n":
            new = False
            data = True
            once = False
            continue
        continue
    if data:
        if line == "\n" and once == False:
            once = True
            continue
        if once and line == "\n":
            continue
        if header == "":
            header = line
        else:
            if not line.count(" selected") and not line.count("old:") and not line.count("new:"):
                rows.append(line)
if rows:
    rows.remove(rows[0])
final.append((query, header, rows))
final.remove(final[0])
file.close()

print('Data Parsing complete')
print('Creating Sheet Lists')

sheet1 = []
sheets = []
for statement in final:
    query, header, rows = statement
    if len(rows) == 0:
        sheet1.append([query])
        sheet1.append([header])
        sheet1.append([])
        continue
    sheet = [query.upper()]
    header = [value.strip() for value in header.split(';')]
    sheet.append(header)
    for row in rows:
        row = [value.strip() for value in row.split(';')]
        sheet.append(row)
    sheets.append(sheet)

workbook = {'ZZSheet1': sheet1}
counter = 2
for sheet in sheets:
    name = "Sheet"+str(counter)
    if sheet[0].count("FROM"):
        name = sheet[0].split("FROM")[1].split(" ")[1]
        if name == "":
            name = "Sheet"+str(counter)
    workbook[name] = sheet[:]
    counter += 1

# sort workbook by sheet name
workbook = dict(sorted(workbook.items(), key=lambda x: x[0]))
filename = sys.argv[1].split('.')[0]+'-'+str(int(time.time()))+'.xlsx'

print("Creating and formatting excel sheet")
filename = os.getcwd()+'\\'+filename
# pyexcel.save_book_as(bookdict=workbook, dest_file_name=filename)
# header style : bold, outline, yellow fill, black text
# data style : outline, black text
wb = openpyxl.Workbook()
for sheet in workbook:
    wb.create_sheet(sheet)
    ws = wb[sheet]
    lengths = []
    if sheet == 'ZZSheet1':
        for row in workbook[sheet]:
            ws.append(row)
    else:
        header = workbook[sheet][1]
        lengths = [len(x) for x in header]
        for i in range(len(header)):
            cell = ws.cell(row=1, column=i+1)
            cell.value = header[i]
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                fill_type='solid', start_color='FFFF00', end_color='FFFF00')
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(
                border_style='thin'), top=openpyxl.styles.Side(border_style='thin'), bottom=openpyxl.styles.Side(border_style='thin'))

        # data
        for row in workbook[sheet][2:]:
            for i in range(len(row)):
                if len(row[i]) > lengths[i]:
                    lengths[i] = len(row[i])
                cell = ws.cell(row=workbook[sheet].index(row), column=i+1)
                cell.value = row[i]
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(
                    border_style='thin'), top=openpyxl.styles.Side(border_style='thin'), bottom=openpyxl.styles.Side(border_style='thin'))

        # query
        length = len(header)
        cell = ws.cell(row=1, column=length+2)
        cell.value = workbook[sheet][0]
    for column_cells in ws.columns:
        try:
            new_column_length = lengths[column_cells[0].column-1]
        except:
            new_column_length = max(len(str(cell.value))
                                    for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.5

# remove default sheet
wb.remove(wb['Sheet'])
print('Saving to file '+filename)
wb.save(filename)
print('Full Path to file: '+filename)
print('Saved Successfully')
